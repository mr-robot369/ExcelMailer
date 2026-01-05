import os
from io import BytesIO

from django.core.mail import EmailMessage
from django.core.exceptions import ValidationError

from openpyxl import load_workbook, Workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from .models import Member, EmailLog
from django.contrib import messages

# ==========================================================
# INTERNAL HELPER: notify success (admin OR dashboard)
# ==========================================================
def _notify_success(admin_instance, request, message):
    """
    Sends success message safely:
    - Admin panel → modeladmin.message_user
    - Dashboard → django.contrib.messages
    """
    if admin_instance is not None:
        # Admin panel
        admin_instance.message_user(request, message)
    else:
        # Frontend dashboard
        messages.success(request, message)


# ==========================================================
# SEND PDF VERSION OF EACH SHEET
# ==========================================================
def process_pdf_and_send_emails(admin_instance, request, queryset):

    for excel_file in queryset:
        wb = load_workbook(excel_file.file.path, data_only=True)

        for sheet_name in wb.sheetnames:
            members = Member.objects.filter(
                sheet_name=sheet_name,
                department=excel_file.department,
                created_by=excel_file.uploaded_by
            )

            for member in members:
                try:
                    buffer = BytesIO()
                    doc = SimpleDocTemplate(buffer, pagesize=letter)

                    sheet = wb[sheet_name]
                    data = [row for row in sheet.iter_rows(values_only=True)]

                    table = Table(data)
                    table.setStyle(TableStyle([
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ]))

                    doc.build([table])
                    pdf_content = buffer.getvalue()
                    buffer.close()

                    email = EmailMessage(
                        subject="Your PDF File",
                        body="Please find attached PDF.",
                        # from_email=os.environ.get("EMAIL_FROM"),
                        from_email=None,
                        to=[member.email],
                    )
                    email.attach(f"{sheet_name}.pdf", pdf_content, "application/pdf")
                    email.send()

                    EmailLog.objects.create(
                        sent_by=request.user,
                        recipient_email=member.email,
                        sheet_name=sheet_name,
                        send_type="pdf",
                        status="success"
                    )

                except Exception as e:
                    EmailLog.objects.create(
                        sent_by=request.user,
                        recipient_email=member.email,
                        sheet_name=sheet_name,
                        send_type="pdf",
                        status="failed",
                        error_message=str(e)
                    )

    _notify_success(admin_instance, request, "PDF emails processed.")



# ==========================================================
# SEND EXCEL SHEETS
# ==========================================================
def process_sheet_and_send_emails(admin_instance, request, queryset):

    for excel_file in queryset:
        wb = load_workbook(excel_file.file.path, data_only=True)

        for sheet_name in wb.sheetnames:
            members = Member.objects.filter(
                sheet_name=sheet_name,
                department=excel_file.department,
                created_by=excel_file.uploaded_by
            )

            for member in members:
                try:
                    buffer = BytesIO()
                    new_wb = Workbook()
                    new_ws = new_wb.active

                    source_sheet = wb[sheet_name]
                    for row in source_sheet.iter_rows():
                        for cell in row:
                            new_ws[cell.coordinate] = cell.value

                    new_wb.save(buffer)
                    excel_content = buffer.getvalue()
                    buffer.close()

                    email = EmailMessage(
                        subject="Your Excel Sheet",
                        body="Please find attached Excel file.",
                        # from_email=os.environ.get("EMAIL_FROM"),
                        from_email=None,
                        to=[member.email],
                    )
                    email.attach(
                        f"{sheet_name}.xlsx",
                        excel_content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    email.send()

                    EmailLog.objects.create(
                        sent_by=request.user,
                        recipient_email=member.email,
                        sheet_name=sheet_name,
                        send_type="excel",
                        status="success"
                    )

                except Exception as e:
                    EmailLog.objects.create(
                        sent_by=request.user,
                        recipient_email=member.email,
                        sheet_name=sheet_name,
                        send_type="excel",
                        status="failed",
                        error_message=str(e)
                    )

    _notify_success(admin_instance, request, "Excel emails processed.")
