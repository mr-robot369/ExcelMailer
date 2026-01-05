
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required

from .models import Profile, Member, ExcelFile, Member, EmailLog
from .utils import (
    process_pdf_and_send_emails,
    process_sheet_and_send_emails
)
from openpyxl import load_workbook
from django.http import Http404, HttpResponse
import csv


# --------------------
# HOME PAGE
# --------------------
def home(request):
    if request.user.is_authenticated:
        return redirect("dashboard")
    return render(request, "home.html")


# --------------------
# LOGIN VIEW (Dept Head Only)
# --------------------
def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is None:
            messages.error(request, "Invalid username or password")
            return redirect("login")

        # ❌ Block admin users from frontend login
        if user.is_staff or user.is_superuser:
            messages.error(request, "Admin users must use admin panel")
            return redirect("login")

        # ✅ Allow only Department Heads
        if not hasattr(user, "profile") or user.profile.role != "head":
            messages.error(request, "You are not authorized to access this portal")
            return redirect("login")

        login(request, user)
        return redirect("dashboard")

    return render(request, "login.html")


# --------------------
# LOGOUT
# --------------------
@login_required
def logout_view(request):
    logout(request)
    return redirect("home")


# --------------------
# DASHBOARD (Placeholder)
# --------------------
@login_required
def dashboard(request):
    user = request.user

    members_count = Member.objects.filter(created_by=user).count()
    excel_count = ExcelFile.objects.filter(uploaded_by=user).count()

    context = {
        "members_count": members_count,
        "excel_count": excel_count,
    }
    return render(request, "dashboard.html", context)

# ----------------------------
# MEMBER LIST
# ----------------------------
@login_required
def member_list(request):
    members = Member.objects.filter(created_by=request.user)

    return render(request, "members/member_list.html", {
        "members": members
    })


# ----------------------------
# ADD MEMBER
# ----------------------------
@login_required
def member_add(request):

    if request.method == "POST":
        name = request.POST.get("name")
        email = request.POST.get("email")
        sheet_name = request.POST.get("sheet_name")

        # Basic validation
        if not name or not email or not sheet_name:
            messages.error(request, "All fields are required")
            return redirect("member_add")

        # Prevent duplicate member in same department
        if Member.objects.filter(
            email=email,
            department=request.user.profile.department
        ).exists():
            messages.error(request, "Member with this email already exists")
            return redirect("member_add")

        Member.objects.create(
            name=name,
            email=email,
            sheet_name=sheet_name,
            department=request.user.profile.department,
            created_by=request.user
        )

        messages.success(request, "Member added successfully")
        return redirect("member_list")

    return render(request, "members/member_add.html")


# ----------------------------
# DELETE MEMBER
# ----------------------------
@login_required
def member_delete(request, pk):
    member = get_object_or_404(
        Member,
        pk=pk,
        created_by=request.user
    )

    member.delete()
    messages.success(request, "Member deleted successfully")
    return redirect("member_list")


@login_required
def excel_upload(request):

    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "Please select an Excel file")
            return redirect("excel_upload")

        if not excel_file.name.endswith((".xlsx", ".xls")):
            messages.error(request, "Only Excel files are allowed")
            return redirect("excel_upload")

        ExcelFile.objects.create(
            file=excel_file,
            uploaded_by=request.user,
            department=request.user.profile.department
        )

        messages.success(request, "Excel file uploaded successfully")
        return redirect("excel_upload")

    files = ExcelFile.objects.filter(uploaded_by=request.user)

    return render(request, "excel/excel_upload.html", {
        "files": files
    })


# Excel send
@login_required
def excel_send(request):

    excel_files = ExcelFile.objects.filter(uploaded_by=request.user)

    excel_with_sheets = []

    for excel in excel_files:
        try:
            wb = load_workbook(excel.file.path, data_only=True)
            excel_with_sheets.append({
                "excel": excel,
                "sheets": wb.sheetnames
            })
        except Exception:
            excel_with_sheets.append({
                "excel": excel,
                "sheets": []
            })

    # ---------------- SEND LOGIC ----------------
    if request.method == "POST":
        excel_id = request.POST.get("excel_id")
        send_type = request.POST.get("send_type")

        excel_file = ExcelFile.objects.filter(
            id=excel_id,
            uploaded_by=request.user
        ).first()

        if not excel_file:
            messages.error(request, "Invalid Excel file")
            return redirect("excel_send")

        queryset = ExcelFile.objects.filter(id=excel_file.id)

        if send_type == "pdf":
            process_pdf_and_send_emails(None, request, queryset)
        elif send_type == "excel":
            process_sheet_and_send_emails(None, request, queryset)
        else:
            messages.error(request, "Invalid send option")

        return redirect("excel_send")

    return render(request, "excel/excel_send.html", {
        "excel_with_sheets": excel_with_sheets
    })

# ! Email Logging
@login_required
def email_logs(request):
    logs = EmailLog.objects.filter(
        sent_by=request.user
    ).order_by("-sent_at")

    return render(request, "email_logs.html", {
        "logs": logs
    })

# ! Excel file Preview
@login_required
def excel_preview(request, excel_id, sheet_name):
    excel_file = get_object_or_404(
        ExcelFile,
        id=excel_id,
        uploaded_by=request.user
    )

    try:
        wb = load_workbook(excel_file.file.path, data_only=True)
    except Exception:
        raise Http404("Unable to read Excel file")

    if sheet_name not in wb.sheetnames:
        raise Http404("Sheet not found")

    sheet = wb[sheet_name]

    # Read first 10 rows for preview
    preview_data = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        preview_data.append(row)
        if i >= 9:
            break

    context = {
        "excel_file": excel_file,
        "sheet_name": sheet_name,
        "preview_data": preview_data,
    }

    return render(request, "excel/excel_preview.html", context)


@login_required
def export_email_logs(request):
    """
    Export email logs as CSV.
    Dept Head → only their logs
    """

    logs = EmailLog.objects.filter(
        sent_by=request.user
    ).order_by("-sent_at")

    response = HttpResponse(
        content_type="text/csv"
    )
    response["Content-Disposition"] = 'attachment; filename="email_logs.csv"'

    writer = csv.writer(response)

    # Header row
    writer.writerow([
        "Recipient Email",
        "Sheet Name",
        "Send Type",
        "Status",
        "Error Message",
        "Sent At"
    ])

    # Data rows
    for log in logs:
        writer.writerow([
            log.recipient_email,
            log.sheet_name,
            log.send_type,
            log.status,
            log.error_message,
            log.sent_at.strftime("%Y-%m-%d %H:%M:%S"),
        ])

    return response
